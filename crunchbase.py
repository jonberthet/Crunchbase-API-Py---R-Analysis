
#Get into right directory
cd ./Desktop/Python/Raw

#get into ipython
ipython --pylab

#Define numpy, pandas, and matplotlib
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt

#Read the Crunchbase API as a .csv file (from their email) and remove commas in the numbers so Python sees the numbers as numeric
df = pd.read_csv('Investment.csv', thousands = ',')

#Just FYI -  List the headers
list(df2.columns.values)

#Remove rows with NA values
df2 = df.dropna()

#For conveneience, rename columns company_category_code and raised_amount_usd to catCode and rAmount, respectively
df3 = df2.rename(columns={'company_category_code' : 'catCode', ' raised_amount_usd ' : 'rAmount'})

#Get sum of each company category.
#Follows this syntax: grouped = df['data1'].groupby(df['key1'])
grouped = df3['rAmount'].groupby(df5['catCode'])
grouped
grouped.sum()

#Plot results below
done = grouped.sum()
done.plot(kind = 'barh', rot=0)


#Rename titles of cbase data:
df = df.rename(columns={'company_category_code' : 'catCode', ' raised_amount_usd ' : 'rAmount', 'funded_year' : 'fundYear', 'investor_country_code' : 'invCountry', 'company_country_code' : 'coCountryCode'})
#Create new dataframe with desired columns (so when I delete NaN rows, I don't delete too many cuz other columns have NaN)
df2 = df[['catCode', 'rAmount', 'coCountryCode']]
#Drop Na's (dropped 10990 rows from 81234 to 70244)
df2 = df2.dropna()
#Group (or Sort) columns in right order
df3 = df2.groupby(['coCountryCode', 'catCode']).sum()
#plot bar graph 
df3.plot(kind = 'barh', rot=0)


####FIX: Strangely, column headers disappear once I use groupby.
#Still graph is too big. Compare 2 countries (US & France)
states = np.array(['USA', 'USA', 'USA', 'ARG', 'ARG', 'ARG'])
category = np.array(['advertising', 'mobile', 'games_video','advertising', 'mobile', 'games_video'])
df2['rAmount'].groupby([states, category]).sum()
df2.groupby(['coCountryCode', 'catCode']).sum() #follows df['datdf.groupby(['key1', 'key2']).mean()
df.groupby(['key1', 'key2']).size()

#separate dataframe into diff. sets, organized by 1 column. Very comprehensive of sorting data
for name, group in df2.groupby('coCountryCode'):
    print name
    print group

#Sort data using multiple columns.
for (k1, k2), group in df2.groupby(['coCountryCode', 'catCode']):
    print k1, k2
    print group

#Gather all rows of a value in a column (ie: all mobile investments)
pieces = dict(list(df2.groupby('catCode')))
pieces['mobile']

#Column-wise and Multiple Function Application (p. 262)
#Import tips file
tips = pd.read_csv('tips.csv')

#create new column called tip_pct, which is tip % of total bill
tips['tip_pct'] = tips['tip'] / tips['total_bill']

#Define a variable
def peak(arr):
   .....:     return arr.max() - arr.min()
grouped.agg(peak)

#Group by diff. columns sex/smoker
grouped = tips.groupby(['sex', 'smoker'])
#Pass name of fn as a string
grouped_pct = grouped['tip_pct']
#Calculate mean of tip_pct
grouped_pct.agg('mean')
#Give me std too!
grouped_pct.agg(['mean', 'std', peak])
####COOOL: Can make a fn apply to a list of columns or specific ones
#Here, I apply count, mean, and max fns to tip_pct and total_bill column
functions = ['count', 'mean', 'max']
result = grouped['tip_pct', 'total_bill'].agg(functions)
result
#If I just want to show the functions done to one column
result['tip_pct']
#Can also change names of columns (shown in book)

#Apply diff. functions to diff. columns
#style is -    column : function
grouped.agg({'tip' : np.max, 'size' : 'sum'})

#Give many functions to tip_pct column but 1 function to size column
grouped.agg({'tip_pct' : ['min', 'max', 'mean', 'std'], 'size' : 'sum'}) 

#Return Data in 'unindexed' form
tips.groupby(['sex', 'smoker'], as_index=False).mean()

#APPLY fn
tips.groupby(['smoker', 'day']).apply(top, n=1, column='total_bill')

#Run fn on certain columns
result = tips.groupby('smoker')['tip_pct'].describe()
result


#Fill NA (NAN) spots values (p. 270)
#Random Sampling, correlation, linear (p. 271)


#Sum up columns together and organize them by dicts (p. 257)
########### 6/5/14 Analysis  ##################
1. What have been the hottest investments in each decade?
2. What industry gets the most funding at what fundraising stage?
3. What regions invest in what industry the most?

#Out of Investment.csv file, rename and make new df
df = df.rename(columns={'company_category_code' : 'catCode', ' raised_amount_usd ' : 'rAmount', 'funded_year' : 'fundYear', 'investor_country_code' : 'invCountry', 'company_country_code' : 'coCountryCode'})
df1 = df[['catCode', 'invCountry', 'fundYear', 'rAmount']]

#Remove NaN values - reduces from 81,234 rows to 55,480 rows
df1 = df1.dropna()
df1

#Organize each catCode by their year (first, organizes by fund yr. Second, organizes by catCode)
for (k1, k2), group in df1.groupby(['fundYear', 'catCode']):
	print k1, k2
	print group

#Organize data by catCode, then fundYear amount (p. 257)
df2 = df1.groupby(['catCode', 'fundYear'])[['rAmount']].sum()
df2.plot(kind = barh)

#Remove organizing
tips.groupby('smoker', group_keys=False).apply(top)

#Fill NA (NAN) spots values (p. 270)
#Random Sampling, correlation, linear (p. 271)

#SAME THING AS ABOVE, but Pivot Tables & Cross-Tabulation
tips.pivot_table(rows=['sex', 'smoker'])

#aggregate only tip_pct and size, and additionally group by day. I’ll put smoker in the table columns and day in the rows:
tips.pivot_table(['tip_pct', 'size'], rows=['sex', 'day'], cols = 'smoker')
#Add mean to all columns at last row called ALL
tips.pivot_table(['tip_pct', 'size'], rows=['sex', 'day'], cols = 'smoker', margins = True)
#'count' or len gives me count or frequency and fill_value puts 0 in empty or Na areas
tips.pivot_table(['tip_pct'], rows=['sex', 'smoker'], cols = 'day', aggfunc = len, margins = True, fill_value = 0)



#####################
####### Scrapy ######
#####################
#Start Here: http://doc.scrapy.org/en/latest/intro/tutorial.html
#Check Version of Python. Must be Python 2.7 or later
python --Version

#Get lxml
pip install lxml

#get help
pip help

#List items pip has installed
pip list

#Show hidden folders
ls -a

#See available commands
scrpay -h

#Open file in terminal (http://hints.macworld.com/article.php?story=2004012218171997)
#open __file__
open pip.log

#Open file in a certain directory, here under Documents
subl ~/Documents/mySite

#Download Scrapy
sudo easy_install Scrapy

#Know present working directory
pwd

#Failed Project
#Create project
scrapy startproject cbase

#Created file in Desktop/Gdrive/ACode/Python/WebCrawl
touch cbase_spider.py

#& Copy/Paste following code inside of it
import scrapy

class CbaseSpider(scrapy.Spider):
    name = "cbase"
    allowed_domains = ["crunchbase.com"]
    start_urls = [
        "http://info.crunchbase.com/about/crunchbase-data-exports/"
    ]

    def parse(self, response):
        filename = response.url.split("/")[-2]
        with open(filename, 'wb') as f:
            f.write(response.body)ls

#Then, go to top level directory and run:
sudo easy_install service_identity

#Then run code below. This creates 2 new files that are the webpages asked for.
scrapy crawl cbase

#Now, I need to 'Select' certain parts w/in the page. I need to choose 'Selectors'
#Start a shell - go to top level directory and run:
scrapy shell "http://info.crunchbase.com/about/crunchbase-data-exports/"

#Extracting Data
#Select each <li> element beloning to the sites list
sel.xpath('//ul/li')

#Get sites descriptions
sel.xpath('//ul/li/text()').extract()

#Get sites titles:
sel.xpath('//ul/li/a/text()').extract()

#Get site Links
sel.xpath('//ul/li/a/@href').extract()

#Now crawling the dmoz.org domain again and you’ll see sites being printed in your output
scrapy crawl cbase

#With this in your dmoz_spider.py, you can 'scrapy crawl dmoz'and get json stuff.
import scrapy

from cbase.items import CbaseItem

class CbaseSpider(scrapy.Spider):
    name = "cbase"
    allowed_domains = ["crunchbase.com"]
    start_urls = [
        "http://info.crunchbase.com/about/crunchbase-data-exports/"
    ]

    def parse(self, response):
        for sel in response.xpath('//ul/li'):
            item = CbaseItem()
            item['title'] = sel.xpath('a/text()').extract()
            item['link'] = sel.xpath('a/@href').extract()
            item['desc'] = sel.xpath('text()').extract()
            yield item


#Store the JSON info into items.json
scrapy crawl cbase -o items.json

##List all the spiders
scrapy list

#Fetch a website
#Cuz, I'm not specifying which area is being downloaded, it doesn't collect here:
scrapy fetch --nolog http://info.crunchbase.com/about/crunchbase-data-exports/

#Cuz I specify what to get, it gives me stuff:
scrapy fetch --nolog --headers http://info.crunchbase.com/about/crunchbase-data-exports/

#View URL in browser
scrapy view http://info.crunchbase.com/about/crunchbase-data-exports/


################
################
##### END OF ###
##### CRAWLER ##
################

####GET DATA####
#With R
#Get in right directory
setwd("/Users/jonathanberthet/Desktop/GDrive/ACode/Python/Cbase")

#Determine wd
getwd()

#See what's inside directory
list.files()

#Download Data from Website
fileUrl <- "http://static.crunchbase.com/exports/crunchbase_monthly_export.xlsx?accessType=DOWNLOAD"

#Put data in directory on computer as a csv
download.file(fileUrl,destfile="/Users/jonathanberthet/Desktop/GDrive/ACode/Python/Cbase/cbase8_14.csv",method="curl")
#Or xls
download.file(fileUrl,destfile="/Users/jonathanberthet/Desktop/GDrive/ACode/Python/Cbase/cbase8_14.xls",method="curl")

#Read certain Tab / doesn't work
cData <- read.xlsx2("./Cbase/cbase8_14.xlsx",sheetIndex=4)

#Find Download Date
dateDownloaded <- date()
dateDownloaded

#Read the csv file
df = read.csv("/cbase814.csv")

#Now that file is in csv format, then open with pandas



#PYTHON
 #Go into pylab
ipython --pylab	

#Define numpy, pandas, and matplotlib
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt

#Know Current WD in Python
import os
os.getcwd()

#See what's in cwd
x = os.listdir('/Users/jonathanberthet/Desktop/GDrive/ACode/Python/Cbase')

#Print out in readable way
for i in x:
	print i

#Save csv worksheet tabs into different worksheets
>>> import pylab
>>> from openpyxl import load_workbook
>>> import codecs
>>> import os
>>> dest_filename = 'cbase814.xlsx'
>>> wb = load_workbook(filename = dest_filename, use_iterators = True)
>>> worksheetNames=wb.get_sheet_names()
>>> worksheetNames
#Find name of certain worksheet
>>> worksheetNames[1:2]
>>> ws = wb['Investments']
>>> ws
<Worksheet "Investments">
>>> for row in ws.iter_rows():
...     for cell in row:
...             val = cell.value
...             print val

import pylab
from openpyxl import load_workbook
import codecs
import os

created_date = '20140730'
basedir = '/Users/jonathanberthet/Desktop/GDrive/ACode/Python/Cbase/'
outputFilenameBase = "/tmp/" + 'crunchbase_monthly_export'

username='root'
password='admin'
host    ='localhost'
database='crunch'

dest_filename = 'cbase814.xlsx'

wb = load_workbook(filename = dest_filename, use_iterators = True)

worksheetNames=wb.get_sheet_names()

#worksheetNames=['Companies']

foCommand = codecs.open(outputFilenameBase + "_" + created_date + ".cmd", "w", "utf-8")

for worksheetName in worksheetNames:
    if worksheetName in ['Analysis', 'License', 'Additions']:
        continue
    worksheetName 
    outputFilename= "_" + worksheetName + "_" + created_date
    fo = codecs.open(outputFilenameBase + outputFilename + ".csv", "w", "utf-8")
    ws = wb[worksheetName]
    columnTypes = []
    columnNames = []
    line = 1
    try:
        for row in ws.iter_rows(): # it brings a new method: iter_rows()
            colIndex=0
            outputLine = []
            if line==100000:
                break
            for cell in row:
                val = cell.value
                if line == 1 :
                    columnNames.append(val)
                try:
                    columnType=columnTypes[colIndex]
                except IndexError:
                    columnTypes.append([])
                    columnType=columnTypes[colIndex]
                                    
                valType =val.__class__.__name__
                if valType != 'NoneType':
                    if not valType in columnType:
                        columnType.append(valType)
                colIndex=colIndex+1
                    
                if val is None:
                    val = ''
                if isinstance(val, basestring):
                    val = val            
                else:                 
                    val = str( val)
                outputLine.append( '"' + val + '"')
            try:
                fo.write( '|'.join(outputLine))
                fo.write( "\n")
            except UnicodeDecodeError:
                print outputLine
                break    
            line=line+1
    except:
        print "ERROR : ", worksheetName
    fo.close()



#Open with Investment Worksheet (http://java.dzone.com/articles/reading-excel-spreadsheets) 
#or (http://www.youlikeprogramming.com/2012/03/examples-reading-excel-xls-documents-using-pythons-xlrd/)
#http://www.simplistix.co.uk/presentations/python-excel.pdf

import workbook = xlrd.open_workbook('cbase8_14.xls')

#Get list of sheet names
print workbook.sheet_names()

#Grab specific worksheet
worksheet = workbook.sheet_by_name('Investments')

#Or Grab by Index
wbook = workbook.sheet_by_index(4)





#Get in right directory
cd /Users/jonathanberthet/Desktop/GDrive/ACode/Python/Cbase

#Read Each of the Files, separate by thousands, and separate info from '|' to different columns.
#Rounds
df.r = pd.read_csv('crunchbase_monthly_export_Rounds_20140730.csv', thousands = ',', error_bad_lines = False, sep = '|')
#Acquisitions
df.a = pd.read_csv('crunchbase_monthly_export_Acquisitions_20140730.csv', thousands = ',', error_bad_lines = False, sep = '|')
#Companies
df.c = pd.read_csv('crunchbase_monthly_export_Companies_20140730.csv', thousands = ',', error_bad_lines = False, sep = '|')
#Investments
df.i = pd.read_csv('crunchbase_monthly_export_Investments_20140730.csv', thousands = ',', error_bad_lines = False, sep = '|')
 
#Read top 5 rows of data
df.r.head()

#Round Questions
***1. How many companies received what stage rounds of fundraising each year?

#Rename columns I need: 
df.r = df.r.rename(columns={'company_category_code' : 'catCode', ' raised_amount_usd ' : 'rAmount', 'funded_year' : 'fundYear', 'company_country_code' : 'coCountry',  ' company_category_list' : 'coCatList'})

#Need company, year, Rount


2. What industries have had the most fundraising rounds per country?
3. Rank the top 5 funded industries per year?

#Investment Questions
1. What have been the top 10 investors' top 5 investments per year?


***2. How many companies received what stage rounds of fundraising each year?

#Rename columns I need: 
df.i = df.i.rename(columns={'company_market' : 'catMarket', 'raised_amount_usd' : 'rAmount', 'funded_year' : 'fundYear', 'company_country_code' : 'coCountry',  'company_category_list' : 'coCatList', 'funding_round_type' : 'fundType'})

#Organize each catCode by their year (first, organizes by fund yr. Second, organizes by catCode)
for (k1, k2), group in df.i.groupby(['fundYear', 'catMarket']):
	print k1, k2
	print group















